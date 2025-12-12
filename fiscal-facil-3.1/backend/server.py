from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Header
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr
from typing import Optional, List
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
from utils.auth import verify_password, get_password_hash, create_access_token, decode_token
from utils.brasil_api import consultar_cnpj
from utils.xml_parser import parse_xml_nota

load_dotenv()

# Configuração MongoDB
MONGO_URL = os.getenv("MONGO_URL", "mongodb://localhost:27017")
client = AsyncIOMotorClient(MONGO_URL)
db = client.fiscal_facil

# FastAPI App
app = FastAPI(title="Fiscal Fácil API", version="2.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== SCHEMAS ====================
class UsuarioRegistro(BaseModel):
    nome: str
    email: EmailStr
    senha: str
    telefone: Optional[str] = None

class UsuarioLogin(BaseModel):
    email: EmailStr
    senha: str

class CnaePermitido(BaseModel):
    cnae_codigo: str
    codigo_servico_municipal: str
    descricao: Optional[str] = None

class EmpresaCadastro(BaseModel):
    cnpj: str
    razao_social: str
    nome_fantasia: Optional[str] = None
    regime_tributario: str
    data_abertura: Optional[str] = None
    cnaes_permitidos: List[CnaePermitido]
    aliquota_imposto: Optional[float] = 6.0  # Alíquota efetiva do Simples Nacional (default: 6% - Anexo III)

# ==================== AUTH MIDDLEWARE ====================
async def get_current_user(authorization: Optional[str] = Header(None)):
    from bson import ObjectId
    
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token não fornecido")
    
    token = authorization.replace("Bearer ", "")
    payload = decode_token(token)
    
    if not payload:
        raise HTTPException(status_code=401, detail="Token inválido ou expirado")
    
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    try:
        user = await db.usuarios.find_one({"_id": ObjectId(user_id)})
    except:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    if not user:
        raise HTTPException(status_code=401, detail="Usuário não encontrado")
    
    return user

# ==================== ROTAS DE AUTENTICAÇÃO ====================
@app.post("/api/auth/registro")
async def registrar_usuario(usuario: UsuarioRegistro):
    # Verifica se o email já existe
    existe = await db.usuarios.find_one({"email": usuario.email})
    if existe:
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    
    # Cria o usuário
    usuario_doc = {
        "nome": usuario.nome,
        "email": usuario.email,
        "senha_hash": get_password_hash(usuario.senha),
        "telefone": usuario.telefone,
        "data_criacao": datetime.utcnow().isoformat()
    }
    
    result = await db.usuarios.insert_one(usuario_doc)
    usuario_id = str(result.inserted_id)
    
    # Gera token
    access_token = create_access_token(data={"sub": usuario_id})
    
    return {
        "mensagem": "Usuário cadastrado com sucesso",
        "access_token": access_token,
        "token_type": "bearer",
        "usuario": {
            "id": usuario_id,
            "nome": usuario.nome,
            "email": usuario.email
        }
    }

@app.post("/api/auth/login")
async def login(credenciais: UsuarioLogin):
    # Busca usuário
    usuario = await db.usuarios.find_one({"email": credenciais.email})
    if not usuario:
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    
    # Verifica senha
    if not verify_password(credenciais.senha, usuario["senha_hash"]):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    
    # Gera token
    usuario_id = str(usuario["_id"])
    access_token = create_access_token(data={"sub": usuario_id})
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "usuario": {
            "id": usuario_id,
            "nome": usuario["nome"],
            "email": usuario["email"]
        }
    }

@app.get("/api/auth/me")
async def obter_usuario_atual(current_user: dict = Depends(get_current_user)):
    return {
        "id": str(current_user["_id"]),
        "nome": current_user["nome"],
        "email": current_user["email"],
        "telefone": current_user.get("telefone")
    }

# ==================== ROTAS DE EMPRESAS ====================
@app.get("/api/empresas/consulta/{cnpj}")
async def consultar_cnpj_endpoint(cnpj: str, current_user: dict = Depends(get_current_user)):
    return consultar_cnpj(cnpj)

@app.post("/api/empresas")
async def cadastrar_empresa(empresa: EmpresaCadastro, current_user: dict = Depends(get_current_user)):
    usuario_id = str(current_user["_id"])
    cnpj_limpo = "".join([n for n in empresa.cnpj if n.isdigit()])
    
    # Verifica se já existe
    existe = await db.empresas.find_one({"cnpj": cnpj_limpo})
    if existe:
        raise HTTPException(status_code=400, detail="Empresa já cadastrada")
    
    # Cria a empresa
    empresa_doc = {
        "usuario_id": usuario_id,
        "cnpj": cnpj_limpo,
        "razao_social": empresa.razao_social,
        "nome_fantasia": empresa.nome_fantasia,
        "regime_tributario": empresa.regime_tributario,
        "data_abertura": empresa.data_abertura,
        "cnaes_permitidos": [cnae.dict() for cnae in empresa.cnaes_permitidos],
        "aliquota_imposto": empresa.aliquota_imposto if hasattr(empresa, 'aliquota_imposto') and empresa.aliquota_imposto else 6.0,  # Default: 6% (Anexo III)
        "data_cadastro": datetime.utcnow().isoformat()
    }
    
    result = await db.empresas.insert_one(empresa_doc)
    
    return {
        "mensagem": "Empresa cadastrada com sucesso",
        "id": str(result.inserted_id)
    }

@app.get("/api/empresas")
async def listar_empresas(current_user: dict = Depends(get_current_user)):
    usuario_id = str(current_user["_id"])
    empresas = []
    
    async for empresa in db.empresas.find({"usuario_id": usuario_id}):
        empresas.append({
            "id": str(empresa["_id"]),
            "cnpj": empresa["cnpj"],
            "razao_social": empresa["razao_social"],
            "nome_fantasia": empresa.get("nome_fantasia"),
            "regime_tributario": empresa["regime_tributario"]
        })
    
    return empresas

@app.get("/api/empresas/{empresa_id}")
async def obter_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    from bson import ObjectId
    
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID inválido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    empresa["id"] = str(empresa.pop("_id"))
    return empresa

# ==================== ROTAS DE NOTAS FISCAIS ====================

# Função auxiliar para processar um único XML
async def processar_xml_nota(empresa_id: str, empresa: dict, conteudo: bytes, nome_arquivo: str):
    """
    Processa um arquivo XML e retorna o resultado do processamento.
    Não levanta exceções, retorna dict com sucesso ou erro.
    """
    try:
        # Parse do XML
        dados_xml = parse_xml_nota(conteudo)
        
        if "erro" in dados_xml:
            return {
                "sucesso": False,
                "nome_arquivo": nome_arquivo,
                "erro": dados_xml["erro"]
            }
        
        # Auditoria: Verifica se o código de serviço está permitido
        codigo_servico = dados_xml['codigo_servico']
        cnaes_permitidos = empresa.get("cnaes_permitidos", [])
        
        cnae_encontrado = None
        for cnae in cnaes_permitidos:
            if cnae.get("codigo_servico_municipal") == codigo_servico:
                cnae_encontrado = cnae
                break
        
        status = "APROVADA"
        mensagem = "Nota fiscal em conformidade"
        
        if not cnae_encontrado:
            status = "ERRO_CNAE"
            mensagem = f"Código de serviço '{codigo_servico}' não autorizado para este CNPJ"
        
        # Salva a nota (incluindo o XML original)
        nota_doc = {
            "empresa_id": empresa_id,
            "numero_nota": dados_xml['numero_nota'],
            "data_emissao": dados_xml['data_emissao'],
            "chave_validacao": dados_xml.get('chave_validacao'),
            "cnpj_tomador": dados_xml.get('cnpj_tomador'),
            "codigo_servico_utilizado": codigo_servico,
            "valor_total": dados_xml['valor_total'],
            "status_auditoria": status,
            "mensagem_erro": mensagem,
            "xml_original": dados_xml.get('xml_bruto', ''),  # Armazena o XML completo
            "data_importacao": datetime.utcnow().isoformat()
        }
        
        result = await db.notas_fiscais.insert_one(nota_doc)
        
        return {
            "sucesso": True,
            "nome_arquivo": nome_arquivo,
            "nota": {
                "id": str(result.inserted_id),
                "numero_nota": nota_doc["numero_nota"],
                "status_auditoria": nota_doc["status_auditoria"],
                "valor_total": nota_doc["valor_total"]
            }
        }
        
    except Exception as e:
        return {
            "sucesso": False,
            "nome_arquivo": nome_arquivo,
            "erro": f"Erro ao processar: {str(e)}"
        }

def normalizar_cnpj(cnpj: str) -> str:
    """Remove caracteres especiais do CNPJ, mantendo apenas dígitos"""
    if not cnpj:
        return ""
    import re
    return re.sub(r'[^0-9]', '', cnpj)

@app.post("/api/notas/importar/{empresa_id}")
async def importar_nota_xml(
    empresa_id: str, 
    file: UploadFile = File(...), 
    current_user: dict = Depends(get_current_user)
):
    from bson import ObjectId
    
    # Verifica se a empresa existe e pertence ao usuário
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inválido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Lê o XML
    conteudo = await file.read()
    
    # Parse do XML para verificar CNPJ ANTES de processar
    dados_xml = parse_xml_nota(conteudo)
    
    if "erro" in dados_xml:
        raise HTTPException(status_code=400, detail=dados_xml["erro"])
    
    # 🔒 VALIDAÇÃO CRÍTICA: Isolamento Multi-Tenant (CNPJ do XML deve bater com CNPJ da Empresa)
    cnpj_empresa_bd = normalizar_cnpj(empresa.get("cnpj", ""))
    cnpj_xml_prestador = dados_xml.get('cnpj_prestador')
    formato_xml = dados_xml.get('formato_xml', 'Desconhecido')
    
    # 🚨 FAIL FAST: Se não conseguiu extrair CNPJ do Prestador, REJEITA por segurança
    if not cnpj_xml_prestador:
        raise HTTPException(
            status_code=400,
            detail=f"🚫 ERRO DE SEGURANÇA: Não foi possível identificar o CNPJ do Prestador neste XML ({formato_xml}). "
                   f"Por questões de segurança e isolamento multi-tenant, este arquivo não pode ser importado. "
                   f"Utilize XMLs no formato SPED (Padrão Nacional) que possuem identificação completa do emissor."
        )
    
    # Normaliza e compara CNPJs
    cnpj_xml_normalizado = normalizar_cnpj(cnpj_xml_prestador)
    
    if cnpj_xml_normalizado != cnpj_empresa_bd:
        # BLOQUEIA IMPORTAÇÃO
        raise HTTPException(
            status_code=400,
            detail=f"🚫 ERRO DE SEGURANÇA: Este XML ({formato_xml}) pertence ao CNPJ {cnpj_xml_prestador}, "
                   f"mas você está tentando importar na empresa {empresa.get('razao_social')} "
                   f"(CNPJ {empresa.get('cnpj')}). Selecione a empresa correta antes de importar."
        )
    
    # Processa o XML normalmente (já passou na validação)
    resultado = await processar_xml_nota(empresa_id, empresa, conteudo, file.filename)
    
    if not resultado["sucesso"]:
        raise HTTPException(status_code=400, detail=resultado["erro"])
    
    # Busca a nota completa para retornar
    from bson import ObjectId
    nota = await db.notas_fiscais.find_one({"_id": ObjectId(resultado["nota"]["id"])})
    
    return {
        "id": str(nota["_id"]),
        "numero_nota": nota["numero_nota"],
        "data_emissao": nota["data_emissao"],
        "codigo_servico_utilizado": nota["codigo_servico_utilizado"],
        "valor_total": nota["valor_total"],
        "status_auditoria": nota["status_auditoria"],
        "mensagem_erro": nota["mensagem_erro"],
        "chave_validacao": nota.get("chave_validacao"),
        "cnpj_tomador": nota.get("cnpj_tomador"),
        "data_importacao": nota["data_importacao"]
    }

@app.post("/api/notas/importar-lote/{empresa_id}")
async def importar_notas_em_lote(
    empresa_id: str,
    files: List[UploadFile] = File(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Importa múltiplos arquivos XML de notas fiscais de uma vez.
    Processa todos os arquivos de forma "graceful" - se um falhar, continua os outros.
    """
    from bson import ObjectId
    
    # Verifica se a empresa existe e pertence ao usuário
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inválido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Valida quantidade de arquivos (máximo 100 por upload)
    if len(files) > 100:
        raise HTTPException(status_code=400, detail="Máximo de 100 arquivos por upload")
    
    # Processa cada arquivo
    resultados = []
    sucessos = 0
    falhas = 0
    detalhes_falhas = []
    
    for idx, file in enumerate(files, 1):
        # Lê o conteúdo
        conteudo = await file.read()
        
        # Processa o XML
        resultado = await processar_xml_nota(empresa_id, empresa, conteudo, file.filename)
        
        if resultado["sucesso"]:
            sucessos += 1
        else:
            falhas += 1
            detalhes_falhas.append({
                "arquivo": file.filename,
                "erro": resultado["erro"]
            })
        
        resultados.append(resultado)
    
    # Retorna resumo
    return {
        "total_arquivos": len(files),
        "sucesso": sucessos,
        "falhas": falhas,
        "detalhes_falhas": detalhes_falhas,
        "resultados": resultados
    }

@app.get("/api/notas/{nota_id}/detalhes")
async def obter_detalhes_nota(nota_id: str, current_user: dict = Depends(get_current_user)):
    """
    Retorna todos os detalhes de uma nota fiscal, incluindo XML original.
    """
    from bson import ObjectId
    
    # Busca a nota
    try:
        nota = await db.notas_fiscais.find_one({"_id": ObjectId(nota_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de nota inválido")
    
    if not nota:
        raise HTTPException(status_code=404, detail="Nota não encontrada")
    
    # Verifica se a empresa da nota pertence ao usuário
    empresa_id = nota.get("empresa_id")
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        pass
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Retorna nota com todos os dados
    return {
        "id": str(nota["_id"]),
        "numero_nota": nota.get("numero_nota"),
        "data_emissao": nota.get("data_emissao"),
        "chave_validacao": nota.get("chave_validacao"),
        "cnpj_tomador": nota.get("cnpj_tomador"),
        "codigo_servico_utilizado": nota.get("codigo_servico_utilizado"),
        "valor_total": nota.get("valor_total"),
        "status_auditoria": nota.get("status_auditoria"),
        "mensagem_erro": nota.get("mensagem_erro"),
        "xml_original": nota.get("xml_original", ''),
        "data_importacao": nota.get("data_importacao"),
        "empresa": {
            "razao_social": empresa.get("razao_social"),
            "cnpj": empresa.get("cnpj"),
            "regime_tributario": empresa.get("regime_tributario")
        }
    }

@app.get("/api/notas/empresa/{empresa_id}")
async def listar_notas_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    from bson import ObjectId
    
    # Verifica se a empresa pertence ao usuário
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inválido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Lista as notas com cálculo de imposto estimado
    notas = []
    async for nota in db.notas_fiscais.find({"empresa_id": empresa_id}):
        valor_total = nota.get("valor_total", 0)
        
        # Cálculo de imposto estimado (Anexo III - 6%)
        imposto_estimado = valor_total * 0.06
        
        notas.append({
            "id": str(nota["_id"]),
            "numero_nota": nota["numero_nota"],
            "data_emissao": nota["data_emissao"],
            "codigo_servico_utilizado": nota["codigo_servico_utilizado"],
            "valor_total": valor_total,
            "imposto_estimado": round(imposto_estimado, 2),  # NOVO
            "status_auditoria": nota["status_auditoria"],
            "mensagem_erro": nota.get("mensagem_erro"),
            "data_importacao": nota["data_importacao"]
        })
    
    return notas

@app.get("/api/notas/estatisticas/{empresa_id}")
async def obter_estatisticas(empresa_id: str, current_user: dict = Depends(get_current_user)):
    from bson import ObjectId
    
    # Verifica acesso
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID inválido")
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Estatísticas
    total = await db.notas_fiscais.count_documents({"empresa_id": empresa_id})
    aprovadas = await db.notas_fiscais.count_documents({"empresa_id": empresa_id, "status_auditoria": "APROVADA"})
    erros = await db.notas_fiscais.count_documents({"empresa_id": empresa_id, "status_auditoria": {"$ne": "APROVADA"}})
    
    # Valor total
    pipeline = [
        {"$match": {"empresa_id": empresa_id}},
        {"$group": {"_id": None, "total": {"$sum": "$valor_total"}}}
    ]
    
    resultado = await db.notas_fiscais.aggregate(pipeline).to_list(1)
    valor_total = resultado[0]["total"] if resultado else 0
    
    # Cálculo de imposto estimado total (Anexo III - 6%)
    imposto_estimado_total = valor_total * 0.06
    
    return {
        "total_notas": total,
        "aprovadas": aprovadas,
        "com_erros": erros,
        "valor_total": valor_total,
        "imposto_estimado_total": round(imposto_estimado_total, 2)  # NOVO
    }

@app.get("/api/notas/imposto-mes/{empresa_id}")
async def obter_imposto_mes(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """
    Retorna o imposto estimado das notas do mês atual.
    Regra: Anexo III do Simples Nacional (6% fixo para MVP).
    """
    from bson import ObjectId
    from dateutil.relativedelta import relativedelta
    
    # Verifica acesso
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID inválido")
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Define o intervalo do mês atual
    hoje = datetime.utcnow()
    primeiro_dia_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    proximo_mes = primeiro_dia_mes + relativedelta(months=1)
    
    # Agregação para somar valores do mês
    pipeline = [
        {
            "$match": {
                "empresa_id": empresa_id,
                "data_emissao": {
                    "$gte": primeiro_dia_mes.isoformat(),
                    "$lt": proximo_mes.isoformat()
                }
            }
        },
        {
            "$group": {
                "_id": None,
                "valor_total_mes": {"$sum": "$valor_total"}
            }
        }
    ]
    
    resultado = await db.notas_fiscais.aggregate(pipeline).to_list(1)
    valor_total_mes = resultado[0]["valor_total_mes"] if resultado else 0.0
    
    # Obtém a alíquota configurada da empresa (default: 6.0%)
    aliquota_empresa = empresa.get("aliquota_imposto", 6.0)
    
    # Cálculo do imposto estimado usando alíquota configurada
    imposto_estimado_mes = valor_total_mes * (aliquota_empresa / 100)
    
    # Determina a base de cálculo baseada no regime
    regime = empresa.get("regime_tributario", "Simples Nacional")
    base_calculo = f"{regime} - Alíquota Configurada"
    
    return {
        "mes_referencia": hoje.strftime('%m/%Y'),
        "valor_total_mes": round(valor_total_mes, 2),
        "imposto_estimado_mes": round(imposto_estimado_mes, 2),
        "aliquota_aplicada": aliquota_empresa,
        "base_calculo": base_calculo
    }

@app.delete("/api/notas/{nota_id}")
async def excluir_nota(nota_id: str, current_user: dict = Depends(get_current_user)):
    """
    Exclui uma nota fiscal. Verifica se a nota pertence a uma empresa do usuário.
    """
    from bson import ObjectId
    
    # Busca a nota
    try:
        nota = await db.notas_fiscais.find_one({"_id": ObjectId(nota_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de nota inválido")
    
    if not nota:
        raise HTTPException(status_code=404, detail="Nota não encontrada")
    
    # Verifica se a empresa da nota pertence ao usuário
    empresa_id = nota.get("empresa_id")
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        pass
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Exclui a nota
    await db.notas_fiscais.delete_one({"_id": ObjectId(nota_id)})
    
    return {
        "mensagem": "Nota excluída com sucesso",
        "nota_id": nota_id
    }

@app.put("/api/empresas/{empresa_id}")
async def atualizar_empresa(
    empresa_id: str,
    dados: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Atualiza dados de uma empresa. Permite editar razão social, nome fantasia,
    regime tributário e CNAEs permitidos.
    """
    from bson import ObjectId
    
    # Verifica se a empresa existe e pertence ao usuário
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inválido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Campos que podem ser atualizados
    campos_permitidos = [
        "razao_social",
        "nome_fantasia",
        "regime_tributario",
        "cnaes_permitidos",
        "limite_faturamento_anual",
        "aliquota_imposto"  # Alíquota efetiva do Simples Nacional
    ]
    
    # Prepara o update
    update_data = {}
    for campo in campos_permitidos:
        if campo in dados:
            # Validação especial para alíquota
            if campo == "aliquota_imposto":
                aliquota = float(dados[campo])
                if aliquota < 0.01 or aliquota > 20.0:
                    raise HTTPException(
                        status_code=400, 
                        detail="Alíquota deve estar entre 0.01% e 20.0%"
                    )
            update_data[campo] = dados[campo]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum campo válido para atualizar")
    
    # Atualiza no banco
    await db.empresas.update_one(
        {"_id": ObjectId(empresa_id)},
        {"$set": update_data}
    )
    
    # Retorna empresa atualizada
    empresa_atualizada = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    empresa_atualizada["id"] = str(empresa_atualizada.pop("_id"))
    
    return {
        "mensagem": "Empresa atualizada com sucesso",
        "empresa": empresa_atualizada
    }

@app.delete("/api/empresas/{empresa_id}")
async def excluir_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """
    Exclui uma empresa e todas as suas notas fiscais associadas.
    """
    from bson import ObjectId
    
    # Verifica se a empresa existe e pertence ao usuário
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inválido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Conta quantas notas serão excluídas
    total_notas = await db.notas_fiscais.count_documents({"empresa_id": empresa_id})
    
    # Exclui todas as notas da empresa
    await db.notas_fiscais.delete_many({"empresa_id": empresa_id})
    
    # Exclui a empresa
    await db.empresas.delete_one({"_id": ObjectId(empresa_id)})
    
    return {
        "mensagem": "Empresa e suas notas excluídas com sucesso",
        "empresa_id": empresa_id,
        "notas_excluidas": total_notas
    }

# ==================== RELATÓRIOS ====================
@app.get("/api/notas/{nota_id}/pdf")
async def gerar_pdf_nota(nota_id: str, current_user: dict = Depends(get_current_user)):
    """
    Gera um PDF formatado da nota fiscal para visualização.
    """
    from bson import ObjectId
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Busca a nota
    try:
        nota = await db.notas_fiscais.find_one({"_id": ObjectId(nota_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de nota inválido")
    
    if not nota:
        raise HTTPException(status_code=404, detail="Nota não encontrada")
    
    # Verifica se a empresa da nota pertence ao usuário
    empresa_id = nota.get("empresa_id")
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        pass
    
    if not empresa or str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Cria o PDF em memória
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=10
    )
    
    # Título
    elements.append(Paragraph("NOTA FISCAL DE SERVIÇOS ELETRÔNICA", title_style))
    elements.append(Spacer(1, 10*mm))
    
    # Informações da Empresa Prestadora
    elements.append(Paragraph("DADOS DA EMPRESA PRESTADORA", header_style))
    
    empresa_data = [
        ['Razão Social:', empresa.get('razao_social', 'N/A')],
        ['CNPJ:', empresa.get('cnpj', 'N/A')],
        ['Regime Tributário:', empresa.get('regime_tributario', 'N/A')]
    ]
    
    empresa_table = Table(empresa_data, colWidths=[40*mm, 130*mm])
    empresa_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(empresa_table)
    elements.append(Spacer(1, 8*mm))
    
    # Dados da Nota Fiscal
    elements.append(Paragraph("DADOS DA NOTA FISCAL", header_style))
    
    # Formata data
    data_emissao = nota.get('data_emissao', '')
    try:
        data_obj = datetime.fromisoformat(data_emissao.replace('Z', '+00:00'))
        data_formatada = data_obj.strftime('%d/%m/%Y às %H:%M')
    except:
        data_formatada = data_emissao
    
    # Status com cor
    status = nota.get('status_auditoria', 'N/A')
    status_color = colors.green if status == 'APROVADA' else colors.red
    
    nota_data = [
        ['Número da Nota:', str(nota.get('numero_nota', 'N/A'))],
        ['Data de Emissão:', data_formatada],
        ['Chave de Validação:', nota.get('chave_validacao', 'N/A')],
        ['Código de Serviço:', nota.get('codigo_servico_utilizado', 'N/A')],
        ['CNPJ Tomador:', nota.get('cnpj_tomador', 'N/A')],
        ['Valor Total:', f"R$ {nota.get('valor_total', 0):.2f}"]
    ]
    
    nota_table = Table(nota_data, colWidths=[40*mm, 130*mm])
    nota_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(nota_table)
    elements.append(Spacer(1, 8*mm))
    
    # Status da Auditoria
    elements.append(Paragraph("RESULTADO DA AUDITORIA", header_style))
    
    auditoria_data = [
        ['Status:', status],
        ['Resultado:', nota.get('mensagem_erro', 'Nota fiscal em conformidade')]
    ]
    
    auditoria_table = Table(auditoria_data, colWidths=[40*mm, 130*mm])
    auditoria_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#dcfce7') if status == 'APROVADA' else colors.HexColor('#fee2e2')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('TEXTCOLOR', (1, 0), (1, 0), status_color),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(auditoria_table)
    elements.append(Spacer(1, 8*mm))
    
    # Rodapé
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    data_importacao = nota.get('data_importacao', '')
    try:
        data_import_obj = datetime.fromisoformat(data_importacao.replace('Z', '+00:00'))
        data_import_formatada = data_import_obj.strftime('%d/%m/%Y às %H:%M')
    except:
        data_import_formatada = data_importacao
    
    elements.append(Spacer(1, 15*mm))
    elements.append(Paragraph(f"Documento gerado em {datetime.utcnow().strftime('%d/%m/%Y às %H:%M')}", footer_style))
    elements.append(Paragraph(f"Importado em: {data_import_formatada}", footer_style))
    elements.append(Paragraph("Fiscal Fácil - Sistema de Auditoria Fiscal", footer_style))
    
    # Gera o PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Retorna como streaming response
    filename = f"nota_{nota.get('numero_nota', 'fiscal')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )

@app.get("/api/relatorios/inconsistencias/{empresa_id}")
async def gerar_relatorio_inconsistencias(empresa_id: str, current_user: dict = Depends(get_current_user)):
    """
    Gera um relatório Excel com todas as notas que possuem inconsistências/erros.
    """
    from bson import ObjectId
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Verifica se a empresa pertence ao usuário
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inválido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Busca apenas notas com erros
    notas_com_erro = []
    async for nota in db.notas_fiscais.find({
        "empresa_id": empresa_id,
        "status_auditoria": {"$ne": "APROVADA"}
    }).sort("data_emissao", -1):
        notas_com_erro.append(nota)
    
    if not notas_com_erro:
        raise HTTPException(
            status_code=404, 
            detail="Nenhuma inconsistência encontrada. Todas as notas estão aprovadas!"
        )
    
    # Cria o Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Inconsistências"
    
    # Cabeçalho do relatório
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Relatório de Inconsistências - {empresa.get('razao_social', '')}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"CNPJ: {empresa.get('cnpj', '')}"
    ws['A3'] = f"Data do Relatório: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}"
    ws['A4'] = f"Total de Inconsistências: {len(notas_com_erro)}"
    
    # Cabeçalhos da tabela
    headers = ['Número da Nota', 'Data de Emissão', 'Código de Serviço', 'Valor (R$)', 'Status', 'Erro Encontrado']
    header_row = 6
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    for row_idx, nota in enumerate(notas_com_erro, header_row + 1):
        ws.cell(row=row_idx, column=1, value=nota.get('numero_nota'))
        
        # Data de emissão formatada
        data_str = nota.get('data_emissao', '')
        if data_str:
            try:
                data_obj = datetime.fromisoformat(data_str.replace('Z', '+00:00'))
                ws.cell(row=row_idx, column=2, value=data_obj.strftime('%d/%m/%Y'))
            except:
                ws.cell(row=row_idx, column=2, value=data_str)
        
        ws.cell(row=row_idx, column=3, value=nota.get('codigo_servico_utilizado'))
        ws.cell(row=row_idx, column=4, value=nota.get('valor_total'))
        ws.cell(row=row_idx, column=5, value=nota.get('status_auditoria'))
        ws.cell(row=row_idx, column=6, value=nota.get('mensagem_erro'))
        
        # Destaca linha com erro
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 50
    
    # Salva em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Retorna como download
    filename = f"inconsistencias_{empresa.get('cnpj', 'empresa')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== DASHBOARD - MONITOR RBT12 ====================
@app.get("/api/dashboard/metrics/{empresa_id}")
async def obter_metricas_rbt12(empresa_id: str, current_user: dict = Depends(get_current_user)):
    from bson import ObjectId
    from dateutil.relativedelta import relativedelta
    
    # Verifica se a empresa pertence ao usuário
    try:
        empresa = await db.empresas.find_one({"_id": ObjectId(empresa_id)})
    except:
        raise HTTPException(status_code=400, detail="ID de empresa inválido")
    
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    if str(empresa.get("usuario_id")) != str(current_user["_id"]):
        raise HTTPException(status_code=403, detail="Acesso negado")
    
    # Calcula data de 12 meses atrás
    hoje = datetime.utcnow()
    doze_meses_atras = hoje - relativedelta(months=12)
    
    # Pipeline de agregação para somar faturamento dos últimos 12 meses
    pipeline = [
        {
            "$match": {
                "empresa_id": empresa_id,
                "data_emissao": {
                    "$gte": doze_meses_atras.isoformat(),
                    "$lte": hoje.isoformat()
                }
            }
        },
        {
            "$group": {
                "_id": None,
                "faturamento_12_meses": {"$sum": "$valor_total"}
            }
        }
    ]
    
    resultado = await db.notas_fiscais.aggregate(pipeline).to_list(1)
    faturamento_atual = resultado[0]["faturamento_12_meses"] if resultado else 0.0
    
    # Obtem limite da empresa (padrão MEI: R$ 81.000,00)
    limite_anual = 81000.00  # Padrão MEI
    
    # Se houver limite cadastrado, usa ele
    if "limite_faturamento_anual" in empresa:
        limite_anual = float(empresa["limite_faturamento_anual"])
    elif empresa.get("regime_tributario") == "Simples Nacional":
        limite_anual = 4800000.00  # Limite Simples Nacional
    elif empresa.get("regime_tributario") == "Lucro Presumido":
        limite_anual = 78000000.00  # Limite Lucro Presumido
    
    # Calcula percentual de uso
    percentual_uso = (faturamento_atual / limite_anual * 100) if limite_anual > 0 else 0
    
    # Define status baseado no percentual
    if percentual_uso >= 100:
        status = "ESTOUROU"
    elif percentual_uso >= 80:
        status = "ALERTA"
    else:
        status = "OK"
    
    # Calcula quanto falta para o limite
    margem_disponivel = limite_anual - faturamento_atual
    
    return {
        "faturamento_atual": round(faturamento_atual, 2),
        "limite": round(limite_anual, 2),
        "percentual_uso": round(percentual_uso, 2),
        "status": status,
        "margem_disponivel": round(margem_disponivel, 2),
        "regime_tributario": empresa.get("regime_tributario", "MEI"),
        "razao_social": empresa.get("razao_social", "")
    }

# ==================== ROTA HOME ====================
@app.get("/")
async def home():
    return {
        "mensagem": "Fiscal Fácil API v2.0 - Sistema de Auditoria Fiscal 🚀",
        "status": "online"
    }

@app.get("/api/health")
async def health_check():
    try:
        await db.command("ping")
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database error: {str(e)}")
